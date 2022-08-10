from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color

movies2021 = {
    "76 Days": {
        "Género": "tragicomedia",
        "Duración(Minutos)": 117,
        "Calificacion": 100
    },
    "Quo Vadis, Aida?": {
        "Género": "wéstern",
        "Duración(Minutos)": 121.,
        "Calificacion": 100
    },
    "Sin señas particulares": {
        "Género": "Musical",
        "Duración(Minutos)": 140.,
        "Calificacion": 100
    },
    "Slalom": {
        "Género": "Documental",
        "Duración(Minutos)": 110,
        "Calificacion": 99
    },
    "Mayor": {
        "Género": "Drama",
        "Duración(Minutos)": 120,
        "Calificacion": 99
    },
    "Martin Luther King y el FBI": {
        "Género": "Drama",
        "Duración(Minutos)": 71,
        "Calificacion": 99
    },
    "Minari, Historia de mi familia": {
        "Género": "Drama",
        "Duración(Minutos)": 121,
        "Calificacion": 98
    },
    "The Father": {
        "Género": "Documental-Experimental",
        "Duración(Minutos)": 98,
        "Calificacion": 93
    },
    "Los Mitchell contra las maquinas": {
        "Género": "Documental",
        "Duración(Minutos)": 220,
        "Calificacion": 98
    },
    "Shiva Baby": {
        "Género": "Fantastico",
        "Duración(Minutos)": 108.,
        "Calificacion": 98
    },
}


def movies_excel_2021(path):
    wb = Workbook()
    ws = wb.active

    ws.title = "Best Movies 2021"

    headings = ["Nombre"] + list(movies2021["76 Days"].keys())
    ws.append(headings)
    for movie in movies2021:
        name_movies = list(movies2021[movie].values())
        ws.append([movie] + name_movies)

    # styles
    thin = Side(border_style="thin", color="00008000")
    c_row = Color(theme=4, tint=0.5)
    c_col = Color(theme=1, tint=0.5)

    # estilos para las columnas
    for col in range(1, 5):
        column = get_column_letter(col)  # funcion para obtener el numero de columna
        ws[column + '1'].font = Font(bold=True, color="00FFFFFF", size=12)
        ws[column + '1'].alignment = Alignment(horizontal="center", vertical="center")
        ws[column + '1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws[column + '1'].fill = PatternFill(start_color=c_col, fill_type="solid")

    # estilos para las filas
    for rows in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col=4):
        for cell in rows:
            if cell.row != 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color=c_row, fill_type="solid")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)


def movies_excel_2020(path):
    wb = Workbook()
    ws = wb.active

    ws.title = "Best Movies 2021"

    # columnas
    headings = ["Nombre"] + list(movies2021["76 Days"].keys())
    ws.append(headings)
    # ws.move_range("F1:I1", rows=1, cols=1)
    for movie in movies2021:
        name_movies = list(movies2021[movie].values())
        ws.append([movie] + name_movies)

    # styles
    thin = Side(border_style="thin", color="00008000")
    c_row = Color(theme=4, tint=0.5)
    c_col = Color(theme=1, tint=0.5)

    # estilos para las columnas
    for col in range(5, 9):
        column = get_column_letter(col)  # funcion para obtener el numero de columna
        ws[column + '5'].font = Font(bold=True, color="00FFFFFF", size=12)
        ws[column + '5'].alignment = Alignment(horizontal="center", vertical="center")
        ws[column + '5'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws[column + '5'].fill = PatternFill(start_color=c_col, fill_type="solid")

    # estilos para las filas
    for rows in ws.iter_rows(min_row=1, max_row=11, min_col=5, max_col=9):
        for cell in rows:
            if cell.row != 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color=c_row, fill_type="solid")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(path)


if __name__ == '__main__':
    movies_excel_2021("Best-Movies-2021.xlsx")
